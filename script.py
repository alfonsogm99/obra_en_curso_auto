from datetime import date, timedelta
from openpyxl import load_workbook
import openpyxl
import logging
import datetime
import calendar
import zipfile
import shutil
import time
import os
import sys


#
print("""  		    _       _____   ____   _   _   _____    ____   	  _____  __  __
            /\     | |     |  ___| / __ \ | \ | | / ____|  / __ \ 	 / ____||  \/  |
           /  \    | |     | |_	  | |  | ||  \| || (___   | |  | |	| |  __ | |\/| |
          / /\ \   | |     |  _|  | |  | || . \`| \___ \  | |  | |	| | |_ || |  | |
         / ____ \  | |___  | |	  | |__| || |\  | ____) | | |__| |	| |__| || |  | |
        /_/    \_\ |_____| |_|     \____/ |_| \_||_____/   \____/ 	 \_____||_|  |_|

        ----------------------------------------------------------------------------
        
        """)



# log
FILENAME_LOGGING = 'log.txt'

while os.path.exists(FILENAME_LOGGING):
    os.remove("log.txt")

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
console_handler = logging.StreamHandler(sys.stdout)
logger.addHandler(console_handler)
file_handler = logging.FileHandler(FILENAME_LOGGING)
file_handler.setLevel(logging.DEBUG)  
logger.addHandler(file_handler)



# temp
try:
    
    logger.info("Eliminando ficheros temporales......."+"\n")
    
    archivos = ["KVPS001","KVPS002","KVPS003"]

    for temp in archivos:
        if os.path.exists(f"{temp}.xlsx"):
            os.remove(f"{temp}.xlsx")
        else:
            logger.info(f'El fichero "{temp}.xlsx" no existe')
        if os.path.exists(f"{temp}_.xlsx"):
            os.remove(f"{temp}_.xlsx")
        else:
            logger.info(f'El fichero "{temp}_.xlsx" no existe')
            
except:
    print("error 1: temp*** ")

logger.info("\n" + ".................................................." + "\n")
logger.info("// Código creado por Alfonso Gómez Medel @ 2024 //"+"\n")


def taller03aAuto():

    directorio = r"C:\\dir"
    home = r"C:\\dir"

    if os.path.exists(directorio):
        
        os.chdir(directorio)

        try:
            archivos = ["KVPS001.csv","KVPS002.csv","KVPS003.csv"]
        
            for i in archivos:
                shutil.move(i, home)
            
            
            os.chdir(home)
            
        except:
            logger.info("ERROR 5555 :/&'")


#taller03aAuto()

#

if os.path.exists("./KVPS001.csv"):
    print("")
else:
    logger.error(f"Los ficheros '.csv' no existen.")
    time.sleep(5)
    exit(1)

if os.path.exists("./KVPS001.xlsx"):
    logger.error(f"Los ficheros '.xlsx' ya existen. Por favor, inténtelo de nuevo.")
    time.sleep(5)
    exit(1)
else:
    print("")


#

today=calendar.day_name[date.today().weekday()]
date=date.today()

weekdays = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
months = {
    "1":'Enero',
    "2":'Febrero',
    "3":'Marzo',
    "4":'Abril',
    "5":'Mayo',
    "6":'Junio',
    "7":'Julio',
    "8":'Agosto',
    "9":'Septiembre',
    "10":'Octubre',
    "11":'Noviembre',
    "12":'Diciembre'
}

now = time.localtime()
weekday_index = now.tm_wday
nmes = datetime.datetime.now().month
mes = months[str(nmes)]
anio = datetime.datetime.now().year
hoy = date.today().strftime("%d-%m-%Y")

# 
def limpiar_caracteres(texto):
    caracteres_no_validos = '\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0B\x0C\x0E\x0F\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1A\x1B\x1C\x1D\x1E\x1F'
    return ''.join(char for char in texto if char not in caracteres_no_validos)

#
archivos = ["KVPS001","KVPS002","KVPS003"]


# 
for nombre_archivo in archivos:
    
    with open(f'{nombre_archivo}.csv', 'r', encoding='latin-1') as file_in:
        lines = file_in.readlines()


    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row_idx, line in enumerate(lines, start=1):
        cleaned_line = limpiar_caracteres(line)
        columns = cleaned_line.strip().split()
        for col_idx, value in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # 
    workbook.save(f'{nombre_archivo}_.xlsx')
    workbook.close()

#
fantasy_zip = zipfile.ZipFile(f'./{hoy}.zip', 'w')
for folder, subfolders, files in os.walk('.'):
    for file in files:
        if file.endswith('_.xlsx'):
            fantasy_zip.write(os.path.join(folder, file), file, compress_type = zipfile.ZIP_DEFLATED)
fantasy_zip.close()


if not os.path.exists(".\\0. reg. dias" + f"\\{anio}"):
    os.makedirs(".\\0. reg. dias" + f"\\{anio}")
    
if not os.path.exists(".\\0. reg. dias" + f"\\{anio}" + f"\\{nmes}. {mes}"):
    os.makedirs(".\\0. reg. dias" + f"\\{anio}" + f"\\{nmes}. {mes}")

if not os.path.exists(".\\0. reg. dias" + f"\\{anio}" + f"\\{nmes}. {mes}" + f"\\{hoy}"):
    os.makedirs(".\\0. reg. dias" + f"\\{anio}" + f"\\{nmes}. {mes}" + f"\\{hoy}")

#
shutil.move(f"{hoy}.zip", ".\\0. reg. dias" + f"\\{anio}" + f"\\{nmes}. {mes}" + f"\\{hoy}" + f"\\{hoy}.zip")


#
logger.info("\n"+"Los ficheros se han separado en columnas con éxito.")
time.sleep(3)

def agregar_datos(fichero_principal, diccionario_codigos):

    libro_principal = openpyxl.load_workbook(fichero_principal)
    hoja_principal = libro_principal.active

    #
    for codigo, datos_fichero in diccionario_codigos.items():

        libro_datos = openpyxl.load_workbook(datos_fichero)
        hoja_datos = libro_datos.active

        columna_g_valor = None

        for fila_datos in hoja_datos.iter_rows(min_row=2, values_only=True):
            if fila_datos[0] == "Total":

                columna_g_valor = fila_datos[6]
                break

        if columna_g_valor is None:
            
            columna_g_valor = 0
            
            if codigo == 'KVPS001':
                hoja_principal['D4'].value = columna_g_valor
            elif codigo == 'KVPS002':
                hoja_principal['D6'].value = columna_g_valor
            elif codigo == 'KVPS003':
                hoja_principal['D7'].value = columna_g_valor
        
        else:

            if codigo == 'KVPS001':
                hoja_principal['D4'].value = columna_g_valor
            elif codigo == 'KVPS002':
                hoja_principal['D6'].value = columna_g_valor
            elif codigo == 'KVPS003':
                hoja_principal['D7'].value = columna_g_valor
                

    # 
    libro_principal.save('output.xlsx')
    libro_principal.close()



fichero_principal = 'plantilla.xlsx'

diccionario_codigos = {'KVPS001': 'KVPS001.xlsx',
                       'KVPS002': 'KVPS002.xlsx',
                       'KVPS003': 'KVPS003.xlsx'}
agregar_datos(fichero_principal, diccionario_codigos)

def agregar_datos_segundo_paso(fichero_salida_primer_paso, diccionario_codigos):

    libro_salida_primer_paso = openpyxl.load_workbook(fichero_salida_primer_paso)
    hoja_salida_primer_paso = libro_salida_primer_paso.active


    for codigo, datos_fichero in diccionario_codigos.items():

        libro_datos = openpyxl.load_workbook(datos_fichero)
        hoja_datos = libro_datos.active

        fila_total_encontrado_segundo_paso = None


        for indice, fila_datos in enumerate(hoja_datos.iter_rows(min_row=2, values_only=True), start=2):
            if fila_datos[0] == "Total":

                fila_total_encontrado_segundo_paso = indice
                break


        if fila_total_encontrado_segundo_paso is None:

            valor_columna_c_segundo_paso=0
            
            if codigo == 'KVPS001':
                hoja_salida_primer_paso['E4'].value = valor_columna_c_segundo_paso
            elif codigo == 'KVPS002':
                hoja_salida_primer_paso['E5'].value = valor_columna_c_segundo_paso
            elif codigo == 'KVPS003':
                hoja_salida_primer_paso['E6'].value = valor_columna_c_segundo_paso
            
        
        else:
            
            fila_destino_celda_c_segundo_paso = fila_total_encontrado_segundo_paso + 1
            valor_columna_c_segundo_paso = hoja_datos.cell(row=fila_destino_celda_c_segundo_paso, column=3).value

            if codigo == 'KVPS001':
                hoja_salida_primer_paso['E4'].value = valor_columna_c_segundo_paso
            elif codigo == 'KVPS002':
                hoja_salida_primer_paso['E5'].value = valor_columna_c_segundo_paso
            elif codigo == 'KVPS003':
                hoja_salida_primer_paso['E6'].value = valor_columna_c_segundo_paso
            

    libro_salida_primer_paso.save('output.xlsx')
    libro_salida_primer_paso.close()


#
fichero_salida_primer_paso = 'output.xlsx'
diccionario_codigos = {'KVPS001': 'KVPS001_.xlsx', 'KVPS002': 'KVPS002_.xlsx', 'KVPS003': 'KVPS003_.xlsx'}
agregar_datos_segundo_paso(fichero_salida_primer_paso, diccionario_codigos)



def agregar_datos_tercer_paso(fichero_salida_segundo_paso, diccionario_codigos):

    libro_salida_segundo_paso = openpyxl.load_workbook(fichero_salida_segundo_paso)
    hoja_salida_segundo_paso = libro_salida_segundo_paso.active

    for codigo, datos_fichero in diccionario_codigos.items():

        libro_datos = openpyxl.load_workbook(datos_fichero)
        hoja_datos = libro_datos.active

        total_encontrado_tercer_paso = 0
        columna_g_valor_tercer_paso = None


        for fila_datos in hoja_datos.iter_rows(min_row=2, values_only=True):
            if fila_datos[0] == "Total":
                total_encontrado_tercer_paso += 1


                if total_encontrado_tercer_paso == 2:
                    columna_g_valor_tercer_paso = fila_datos[6]
                    break

        
        if columna_g_valor_tercer_paso is not None:
            
            if codigo == 'KVPS001':
                hoja_salida_segundo_paso['F4'].value = columna_g_valor_tercer_paso
            elif codigo == 'KVPS002':
                hoja_salida_segundo_paso['F5'].value = columna_g_valor_tercer_paso
            elif codigo == 'KVPS003':
                hoja_salida_segundo_paso['F6'].value = columna_g_valor_tercer_paso
                

    libro_salida_segundo_paso.save('output.xlsx')
    libro_salida_segundo_paso.close()


# 
fichero_salida_segundo_paso = 'output.xlsx'
diccionario_codigos = {'KVPS001': 'KVPS001_.xlsx', 'KVPS002': 'KVPS002_.xlsx', 'KVPS003': 'KVPS003_.xlsx'}
agregar_datos_tercer_paso(fichero_salida_segundo_paso, diccionario_codigos)




def agregar_datos_cuarto_paso(fichero_salida_tercer_paso, diccionario_codigos):
    
    libro_salida_tercer_paso = openpyxl.load_workbook(fichero_salida_tercer_paso)
    hoja_salida_tercer_paso = libro_salida_tercer_paso.active

    for codigo, datos_fichero in diccionario_codigos.items():
        print(f"\nProcesando código: {codigo}")

        libro_datos = openpyxl.load_workbook(datos_fichero)
        hoja_datos = libro_datos.active


        total_encontrado_cuarto_paso = 0
        columna_c_valor_cuarto_paso = None


        for fila_datos in hoja_datos.iter_rows(min_row=2, values_only=True):
            if fila_datos[0] == "Total":
                total_encontrado_cuarto_paso += 1
                    
                    
                if total_encontrado_cuarto_paso == 2:
                    fila_total_mas_uno = hoja_datos[fila_datos[0].row + 1]
                    columna_c_valor_cuarto_paso = fila_total_mas_uno[2].value
                    break
                
                
        if columna_c_valor_cuarto_paso is not None:
            print(f"Valor de C encontrado (cuarto paso): {columna_c_valor_cuarto_paso}")

            if codigo == 'KVPS001':
                celda_destino = 'G4'
            elif codigo == 'KVPS002':
                celda_destino = 'G5'
            elif codigo == 'KVPS003':
                celda_destino = 'G6'
                
                
            hoja_salida_tercer_paso[celda_destino].value = columna_c_valor_cuarto_paso

    libro_salida_tercer_paso.save('output.xlsx')
    libro_salida_tercer_paso.close()
    

def agregar_datos_cuarto_paso(fichero_salida_tercer_paso, diccionario_codigos):
    
    libro_salida_tercer_paso = openpyxl.load_workbook(fichero_salida_tercer_paso)
    hoja_salida_tercer_paso = libro_salida_tercer_paso.active


    for codigo, datos_fichero in diccionario_codigos.items():

        libro_datos = openpyxl.load_workbook(datos_fichero)
        hoja_datos = libro_datos.active

        total_encontrado_cuarto_paso = 0
        columna_c_valor_cuarto_paso = None


        for indice_fila, fila_datos in enumerate(hoja_datos.iter_rows(min_row=2, values_only=True), start=2):
            if fila_datos[0] == "Total":
                total_encontrado_cuarto_paso += 1

                if total_encontrado_cuarto_paso == 2 and (indice_fila + 1) <= hoja_datos.max_row:
                    columna_c_valor_cuarto_paso = hoja_datos.cell(row=indice_fila + 1, column=3).value
                    break
                
                
        if columna_c_valor_cuarto_paso is not None:

            if codigo == 'KVPS001':
                celda_destino = 'G4'
            elif codigo == 'KVPS002':
                celda_destino = 'G5'
            elif codigo == 'KVPS003':
                celda_destino = 'G6'
            
            hoja_salida_tercer_paso[celda_destino].value = columna_c_valor_cuarto_paso

    libro_salida_tercer_paso.save('output.xlsx')
    libro_salida_tercer_paso.close()

# 
fichero_salida_tercer_paso = 'output.xlsx'
diccionario_codigos = {'KVPS001': 'KVPS001_.xlsx', 'KVPS002': 'KVPS002_.xlsx', 'KVPS003': 'KVPS003_.xlsx'}
agregar_datos_cuarto_paso(fichero_salida_tercer_paso, diccionario_codigos)

def llenar_celdas_vacias_con_cero(hoja):
    
    #EMPRESA01
    rangos = [('F', 4, 13), ('G', 4, 13)]

    for columna, fila_inicio, fila_fin in rangos:
        for fila in range(fila_inicio, fila_fin + 1):
            celda = hoja[f"{columna}{fila}"]
            if celda.value is None:
                celda.value = 0
    
    
    #EMPRESA02
    rangos2 = [('F', 16, 19), ('G', 16, 19)]

    for columna2, fila_inicio2, fila_fin2 in rangos2:
        for fila2 in range(fila_inicio2, fila_fin2 + 1):
            celda2 = hoja[f"{columna2}{fila2}"]
            if celda2.value is None:
                celda2.value = 0
    
    
    #EMPRESA03
    rangos3 = [('F', 22, 22), ('G', 22, 22)]

    for columna3, fila_inicio3, fila_fin3 in rangos3:
        for fila3 in range(fila_inicio3, fila_fin3 + 1):
            celda3 = hoja[f"{columna3}{fila3}"]
            if celda3.value is None:
                celda3.value = 0

#
libro_salida_final = openpyxl.load_workbook('output.xlsx')
hoja_salida_final = libro_salida_final.active


llenar_celdas_vacias_con_cero(hoja_salida_final)


def convertir_a_numeros(hoja):
    for fila in hoja.iter_rows():
        for celda in fila:

            if celda.value is None:
                continue

            if isinstance(celda.value, (int, float)):
                continue

            valor = str(celda.value).strip()

            try:
                # formato: 1.234,56
                if "," in valor:
                    valor = valor.replace(".", "")
                    valor = valor.replace(",", ".")


                valor = valor.replace(" ", "")

                numero = float(valor)

                if numero.is_integer():
                    celda.value = int(numero)
                else:
                    celda.value = numero

            except:
                pass      






hoy = date.today().strftime("%d-%m-%Y")
months = {
    "1":'Enero',
    "2":'Febrero',
    "3":'Marzo',
    "4":'Abril',
    "5":'Mayo',
    "6":'Junio',
    "7":'Julio',
    "8":'Agosto',
    "9":'Septiembre',
    "10":'Octubre',
    "11":'Noviembre',
    "12":'Diciembre'
}

#
archivo_excel0019 = f'Obra en curso Control {hoy}.xlsx'

#
convertir_a_numeros(hoja_salida_final)
libro_salida_final.save(f'{archivo_excel0019}')
libro_salida_final.close()


if os.path.exists('output.xlsx'):
    os.remove('output.xlsx')



# cantidad de or
from datetime import date, timedelta
from openpyxl import load_workbook
import pandas as pd
import openpyxl
import datetime
import calendar
import shutil
import time
import glob
import os


#
logger.info("\n"+"Información añadida a Obra en curso con éxito.")
logger.info("\n"+"Revisando la cantidad de Órdenes de Reparación ...")




# caracteres no válidos
def limpiar_caracteres(texto):
    caracteres_no_validos = '\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0B\x0C\x0E\x0F\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1A\x1B\x1C\x1D\x1E\x1F'
    return ''.join(char for char in texto if char not in caracteres_no_validos)

# 
archivos = ["KVPS001", "KVPS002", "KVPS003"]

for nombre_archivo in archivos:
    
    with open(f'{nombre_archivo}.csv', 'r', encoding='latin-1') as file_in:
        lines = file_in.readlines()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    #
    for row_idx, line in enumerate(lines, start=1):
        cleaned_line = limpiar_caracteres(line)
        columns = cleaned_line.strip().split()
        for col_idx, value in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # 
    workbook.save(f'{nombre_archivo}.xlsx')
    workbook.close()


archivos_a_procesar = ["KVPS001", "KVPS002", "KVPS003"]

sumas_dias = []

for name in archivos_a_procesar:

    datos00 = f"{name}.xlsx"
    libro_datos = load_workbook(filename=datos00)
    hoja_datos = libro_datos.active
    
    suma_dias = 0

    for fila in hoja_datos.iter_rows(values_only=True):
        valor_a = fila[0]
        valor_e = fila[4] if len(fila) > 4 else None

        if valor_a:
            valor_a = str(valor_a).strip()

            if valor_a.startswith(("OA","OB"....)) or valor_a in ("T"):

                if valor_e is not None:
                    try:
                        valor_e = float(valor_e)
                        suma_dias += valor_e
                    except:
                        pass

    sumas_dias.append(suma_dias)


    #
    for col in reversed(range(2, hoja_datos.max_column + 1)):
        hoja_datos.delete_cols(col)

    filas_a_eliminar_temp = []

    for fila in range(1, hoja_datos.max_row + 1):
        valor_celda = hoja_datos[f"A{fila}"].value

        if valor_celda and not (
            valor_celda.startswith(("OA","OB"....)) or valor_celda in ("T")
        ):
            filas_a_eliminar_temp.append(fila)

    for fila in reversed(filas_a_eliminar_temp):
        hoja_datos.delete_rows(fila)

    libro_datos.save(filename=f"{name}.xlsx")
    libro_datos.close()

    for fila_a_eliminar in reversed(filas_a_eliminar_temp):
        hoja_datos.delete_rows(fila_a_eliminar)
        
        

    # os.rename(f"{name}.xlsx",f"backup-{name}.xlsx")
    
    #    
    libro_datos.save(filename=f"{name}.xlsx")
    libro_datos.close()
    

###
    
def contar_columna_a(archivo):
    try:
        
        workbook = load_workbook(filename=archivo, read_only=True)
        
        sheet = workbook.active
        
        contador = 0
        
        for row in sheet.iter_rows(values_only=True):
            valor_columna_a = row[0] 
            if valor_columna_a is not None:
                contador += 1
        
        return contador
    except Exception as e:
        print(f"No se pudo procesar {archivo}: {str(e)}")
        return None
    
    workbook.close()



recuentos = []

archivos = ["KVPS001", "KVPS002", "KVPS003"]

for nombre in archivos:
    try:
        with open(f"{nombre}.csv", "r", encoding="latin-1") as f:
            
            recuento = sum(
                1 for line in f
                if line.startswith(("OA","OB","OC"......))
            )
            recuentos.append(recuento)
    except FileNotFoundError:
        recuentos.append(0)
        

from datetime import date, timedelta
today=calendar.day_name[date.today().weekday()]
hoy_date = date.today()
hoy = hoy_date.strftime("%d-%m-%Y")
anio = datetime.datetime.now().year
months = {
    "1":'Enero',
    "2":'Febrero',
    "3":'Marzo',
    "4":'Abril',
    "5":'Mayo',
    "6":'Junio',
    "7":'Julio',
    "8":'Agosto',
    "9":'Septiembre',
    "10":'Octubre',
    "11":'Noviembre',
    "12":'Diciembre'
}
nmes = datetime.datetime.now().month
mes = months[str(nmes)]

#
archivo_excel009 = f'Obra en curso Control {hoy}.xlsx'

#
workbook009 = openpyxl.load_workbook(archivo_excel009)
hoja009 = workbook009.active  

# 
a = recuentos[0]
b = recuentos[1]
c = recuentos[2]
d = recuentos[3]
e = recuentos[4]
f = recuentos[5]
g = recuentos[6]
h = recuentos[7]
i = recuentos[8]
j = recuentos[9]
k = recuentos[10]
l = recuentos[11]
m = recuentos[12]
n = recuentos[13]
o = recuentos[14]

#
j1 = sumas_dias[0]
j2 = sumas_dias[1]
j3 = sumas_dias[2]
j4 = sumas_dias[3]
j5 = sumas_dias[4]
j6 = sumas_dias[5]
j7 = sumas_dias[6]
j8 = sumas_dias[7]
j9 = sumas_dias[8]
j10 = sumas_dias[9]
j11 = sumas_dias[10]
j12 = sumas_dias[11]
j13 = sumas_dias[12]
j14 = sumas_dias[13]
j15 = sumas_dias[14]

#
hoja009['C4'] = a
hoja009['C5'] = b
hoja009['C6'] = c
hoja009['C7'] = d
hoja009['C8'] = e
hoja009['C9'] = f
hoja009['C10'] = g
hoja009['C11'] = h
hoja009['C12'] = i
hoja009['C13'] = j
hoja009['C16'] = k
hoja009['C17'] = l
hoja009['C18'] = m
hoja009['C19'] = n
hoja009['C22'] = o

#
hoja009['J4'] = j1
hoja009['J5'] = j2
hoja009['J6'] = j3
hoja009['J7'] = j4
hoja009['J8'] = j5
hoja009['J9'] = j6
hoja009['J10'] = j7
hoja009['J11'] = j8
hoja009['J12'] = j9
hoja009['J13'] = j10
hoja009['J16'] = j11
hoja009['J17'] = j12
hoja009['J18'] = j13
hoja009['J19'] = j14
hoja009['J22'] = j15



workbook009.save(archivo_excel009)
workbook009.close()

time.sleep(5)


from datetime import datetime

wb = openpyxl.load_workbook(archivo_excel009)

fecha_actual_formateada = datetime.now().strftime('%d.%m.%Y')

wb.active.title = fecha_actual_formateada
wb.save(archivo_excel009)
wb.close()


time.sleep(2)



#
if not os.path.exists("C:\\dir" + f"\\{anio}"):
    os.makedirs("C:\\dir" + f"\\{anio}")
    
if not os.path.exists("C:\\dir" + f"\\{anio}\\{nmes}. {mes}"):
    os.makedirs("C:\\dir" + f"\\{anio}"+ f"\\{nmes}. {mes}")


time.sleep(2)

shutil.move(archivo_excel009,f"C:\\dir\\{anio}\\{nmes}. {mes}\\{archivo_excel009}")
   
for archivo in archivos:
    shutil.move(f"{archivo}.csv", ".\\0. reg. dias" + f"\\{anio}" + f"\\{nmes}. {mes}" + f"\\{hoy}")


#
logger.info("\n"+"Listo!")
time.sleep(5)